#读写xlsx文件
from openpyxl import load_workbook,Workbook
import easygui as g

def turnStr(L):
    """ 把列表里所有元素转换为字符串型 """
    tmpL = list(map(lambda x: str(x), L))
    return tmpL

#创建xlsx文件
def createWB(excelFile):
    wb = Workbook()
    wb.save(excelFile)
    
#创建sheet表
def createWS(excelFile,sheetname):
    try:
        wb = load_workbook(excelFile)
    except FileNotFoundError as reason:
        print(reason)
        g.msgbox("%s 该文件不存在！"%(excelFile))
    ws = wb.create_sheet(title=sheetname)
    wb.save(excelFile)
    
#删除sheet表
def deleteWS(excelFile,sheetname):
    try:
        wb = load_workbook(excelFile)
    except FileNotFoundError as reason:
        print(reason)
        g.msgbox("%s 该文件不存在！"%(excelFile))
    try:
        ws = wb[sheetname]
        wb.remove(ws)
    except KeyError as reason:
        print(reason)
        pass
    try:
        wb.save(excelFile)
    except IndexError as reason:#至少要保留一张sheet，没有就创建一个同名的
        print(reason)
        ws = wb.create_sheet(title=sheetname)
    wb.save(excelFile)
    
#读取xlsx文件
def readXlsx(excelFile,*args):
    try:
        wb = load_workbook(excelFile)
    except FileNotFoundError as reason:
        print(reason)
        g.msgbox("%s 该文件不存在！"%(excelFile))
    if len(args) > 0:
        sheetname = args[0]
    else:
        sheetname = wb.get_sheet_names()[0]
    ws = wb[sheetname]
    data = []
    column_index = 0#用于定位行数
    for columons in ws.rows:
        data.append([])
        for member in columons:
            data[column_index].append(member.value)
        column_index += 1
    return (data)
    
#写入xlsx文件，以list格式写入到最后一行
def writeXlsx(excelFile,data):
    try:
        wb = load_workbook(excelFile)
    except FileNotFoundError as reason:
        print(reason)
        createXlsx(excelFile)
    ws = wb.active
    ws.append(data)
    wb.save(excelFile)
