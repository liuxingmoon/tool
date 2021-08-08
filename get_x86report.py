#!python3
# --*-- coding: utf-8 --*--

import os,re
import glob
from openpyxl import load_workbook
import xlrd
import pdb
from file_ctrl import *
import threading
import datetime
import remind_write_weekreport as remind
import easygui as g
import svnconfig,svn_ctrl
from remind_write_weekreport import x86_members
import time,configparser
from clip_ctrl import clip
from file_ctrl import replace

configpath = r"Config.ini"
config = configparser.ConfigParser()
config.read(configpath, encoding="utf-8")
weekReportDir = r"%s" %(config.get("work", "weekReportDir"))
weekReportDir_tradition = r"%s" %(config.get("work", "weekReportDir_tradition"))
weekReportDir_cloud = r"%s" %(config.get("work", "weekReportDir_cloud"))
totalFile = r"%s" %(config.get("work", "totalFile"))
weekReport_myself = r"%s" %(config.get("work", "weekReport_myself"))
today_date = datetime.datetime.today().strftime("%Y-%m-%d")#今天日期
reportFile = r"x86组周报_%s.txt"%(today_date)
dist_tradition = svnconfig.setting['dist_tradition']
dist_cloud = svnconfig.setting['dist_cloud']

work_items = {
    "生产值守":"","应急处置":"","故障处理":"","工单处理":"","变更实施":"","容灾准备":"","高可用建设":"","灾备演练":"","咨询答疑":"","巡检平台":"","监控告警":"","配置库":"","蓝鲸平台":"","璇玑系统":"","系统改造":"","DNS改造":"","NAS迁移":"","能力建设":"","流程规范":"","文档编写":"","其他":""
}
    
def turnStr(L):
    """ 把列表里所有元素转换为字符串型 """
    tmpL = list(map(lambda x: str(x), L))
    return tmpL

def readXls(excelFile):
    row_pd = 0
    row_ha = 0
    row_prog = 0
    row_tool = 0
    row_special = 0
    row_other = 0
    row_end = 0
    wb = xlrd.open_workbook(excelFile)
    ws = wb.sheets()[0]
    nrows = ws.nrows
    # 
    data = []
    for row_num in range(1,nrows):
        celldata = ws.cell_value(row_num,1)
        if celldata:
            if celldata == "上周安排工作完成情况":
                data.append(celldata)
                break
            else:
                data.append(celldata)
        else:
            continue
    # 
    for ind,celldata in enumerate(data):
        #print("{} === {}".format(ind, celldata))
        if celldata == "生产运行保障":
            row_pd = ind
        if celldata == "高可用性管理":
            row_ha = ind  
        if celldata == "项目支持":
            row_prog = ind 
        if celldata == "工具建设":
            row_tool = ind
        if celldata == "专项工作":
            row_special = ind
        if celldata == "其他工作":
            row_other = ind
        if celldata == "上周安排工作完成情况":
            row_end = ind
            break

    excelData = [
        data[(row_pd +1):row_ha],
        data[(row_ha +1):row_prog],
        data[(row_prog +1):row_tool],
        data[(row_tool +1):row_special],
        data[(row_special +1):row_other],
        data[(row_other +1):row_end],
    ]
    return excelData


def readXlsx(excelFile):
    row_pd = 0
    row_ha = 0
    row_prog = 0
    row_tool = 0
    row_special = 0
    row_other = 0
    
    wb = load_workbook(excelFile)
    ws = wb.active
    
    for c in ws.rows:
        if c[1].value:
            if c[1].value == "生产运行保障":
                row_pd = c[1].row + 1
            elif c[1].value == "高可用性管理":
                row_ha = c[1].row + 1
            elif c[1].value == "项目支持":
                row_prog = c[1].row + 1
            elif c[1].value == "工具建设":
                row_tool = c[1].row + 1
            elif c[1].value == "专项工作":
                row_special = c[1].row + 1
            elif c[1].value == "其他工作":
                row_other = c[1].row + 1          
            elif c[1].value == "上周安排工作完成情况":
                break
            else:
                continue
        else:
            continue
        
    return [
        ws.cell(row_pd, 2).value.split("\n") if ws.cell(row_pd, 2).value else [],
        ws.cell(row_ha, 2).value.split("\n") if ws.cell(row_ha, 2).value else [],
        ws.cell(row_prog, 2).value.split("\n") if ws.cell(row_prog, 2).value else [],
        ws.cell(row_tool, 2).value.split("\n") if ws.cell(row_tool, 2).value else [],
        ws.cell(row_special, 2).value.split("\n") if ws.cell(row_special, 2).value else [],
        ws.cell(row_other, 2).value.split("\n") if ws.cell(row_other, 2).value else [],
        ]

def clean_weekdir():
    os.chdir(weekReportDir)
    filelist = os.listdir()
    rm_filelist = [x for x in filelist if x != "x86_weekreport.txt"]
    for rm_file in rm_filelist:
        os.remove(rm_file)
        
        
def getfilelist(filedir):
    os.chdir(filedir)
    last_dir = max([ datetime.datetime.strptime( dir_name, '%Y-%m-%d') for dir_name in os.listdir() ])#最新的时间的目录
    today_date = datetime.datetime.today()#今天日期
    if ((today_date - last_dir).days >= 7):#如果当前时间比上周的提交目录已经过了一周，自动创建周报目录并提交
        last_dir = today_date.strftime("%Y-%m-%d")
        os.system("mkdir %s" %(last_dir))
        #提交周报目录
        commit_dir = filedir + os.sep + last_dir
        svn_ctrl.add(commit_dir)
        svn_ctrl.commit(commit_dir)
    else:
        last_dir = last_dir.strftime("%Y-%m-%d")
    os.chdir(last_dir)#进入到最新的目录
    file_list = os.listdir()#周报列表
    return(last_dir,file_list)
    
def copy_file_to_weekdir(filedir):
    last_dir = getfilelist(filedir)[0]
    file_list_temp = getfilelist(filedir)[1]
    file_list = []
    for member in x86_members:
        for file in file_list_temp:
            if member in file:#周报文件在x86member中，就加入到复制列表中
                file_list.append(file)
    file_list = list(set(file_list))#去重
    for filename in file_list:
        thread_copyfile =threading.Thread(target=copy_file,args=(filename,filedir + os.sep + last_dir,weekReportDir))#从当前目录拷贝文件到汇总目录
        thread_copyfile.start()

def get_item(list,member):
    #获取子项的标签，通过递归获取前一行的信息标签来获取
    try:
        #item = re.findall(r"\[.*\]",list[list.index(member) - 1])[0]
        item = re.split("[\[\]]",re.findall(r"\[.*\]",list[list.index(member) - 1])[0])[1]
        #item = item.strip('[]')
        if item in work_items:
            flag_get_item = True
        else:
            flag_get_item = False
            print("%s 不在选项标签中，请检查"%s(item))
    except IndexError as reason:
        flag_get_item = False
        pass
    if flag_get_item:
        print(item,member)
        return (item)
    else:
        get_item(list,list[list.index(member) - 1])
    
if __name__ == "__main__":
    # TODO: 检查周报是否交齐,
    # x86_members = ["任德强","张文强","刁强",]
    #更新svn周报到本地仓库
    svn_ctrl.update(dist_tradition)
    svn_ctrl.update(dist_cloud)
    #提交我的周报
    last_dir = getfilelist(dist_cloud)[0]
    dist_dir = dist_cloud + os.sep + last_dir#目标目录
    locallist = os.listdir(weekReport_myself)
    #检查自己是否已提交周报
    check_day_file = last_dir.replace("-","")#提取周报日
    my_report = [x for x in locallist if '工作周报-刘兴%s' %(check_day_file) in x]
    if my_report != []:
        my_report = my_report[0]
    else:
        print('刘兴的周报还没写！')
        g.msgbox(title='刘兴周报检查',msg='刘兴的周报（工作周报-刘兴%s）还没写！写完后确认' %(check_day_file))
        my_report = '工作周报-刘兴%s' %(check_day_file)
    commited_files = os.listdir(dist_dir)#已提交周报的list
    if [x for x in commited_files if '工作周报-刘兴%s' %(check_day_file) in x] != []:
        print('刘兴已提交，跳过自己上传周报')
    else:
        print('刘兴未提交，自己上传周报')
        copy_file(my_report,weekReport_myself,dist_dir)
        commit_file = dist_dir + os.sep + my_report
        svn_ctrl.add(commit_file)
        svn_ctrl.commit(commit_file)

    #检查下周报结果，如果都交了就统计,没交完就不汇总统计
    result = remind.start()
    if result != "周报已全部提交，可以汇总了":
        g.msgbox(title='周报检查',msg=result)
        clip(result)
    else:#汇总
        work = {"生产运行保障":[], 
            "高可用性管理":[], 
            "项目支持":[],
            "工具建设":[],
            "专项工作": [],
            "其他工作":[]
        }
        
        clean_weekdir()#清空周报文件
        copy_file_to_weekdir(weekReportDir_tradition)#复制周报文件
        time.sleep(1)#多线程复制文件，可能会导致目录切换报错，等待一秒
        copy_file_to_weekdir(weekReportDir_cloud)#复制周报文件
        time.sleep(1)
        os.chdir(weekReportDir)
        files_count = len([x for x in os.listdir() if '工作周报-' in x])#周报文件数量
        files_deal = 0#已汇总周报文件数量
        for excel in glob.glob('工作周报-*.xls*'):
            print(f"正在处理 {excel} ...")
            if os.path.splitext(excel)[-1] == ".xlsx":  
                data = readXlsx(excel)
            elif os.path.splitext(excel)[-1] == ".xls":  
                data = readXls(excel)
            else:
                print(f"{excel} excel格式不支持...")
                continue
                
            if data[0]:
                work["生产运行保障"] += data[0]
            if data[1]:        
                work["高可用性管理"] += data[1]
            if data[2]: 
                work["项目支持"] += data[2]
            if data[3]: 
                work["工具建设"] += data[3]
            if data[4]: 
                work["专项工作"] += data[4]
            if data[5]:
                work["其他工作"] += data[5]
            files_deal += 1
        ignore_keywords = ["","无","暂无","不涉及","本周不涉及","上周安排工作完成情况"] + list(work.keys())
        with open(totalFile, "w") as fw:
            fw.write("农信x86组周报\n")
            for work_type, work_contents in work.items():
                fw.write("\n")
                fw.write("[{}]\n".format(work_type))
                if work_contents:
                    for line in set(turnStr(work_contents)):
                        if line.strip() not in ignore_keywords or "本周主要工作内容如下" in line:
                            fw.write("{}\n".format(line.strip()))
                fw.write("\n")
        print("汇总结束，总共周报文件数量：%s ;\n已汇总周报文件数量： %s" %(files_count,files_deal))
        
        with open(totalFile, "r") as f:#按照标签重新排序
            son_items = []#存储子项
            file = f.readlines()
            for work_item in work_items:
                work_item_flag = False
                task_num = 1
                for line in file:
                    if work_item in line:
                        if work_item_flag == False:#每个item首行打印
                            if work_items[work_item] == "":
                                work_items[work_item] += "[%s]"%work_item+"\n"
                            work_item_flag = True
                        line_re = re.sub(r"\d+、",str(task_num)+"、",line,1)#从左往右替换一次数字
                        line_re = re.sub(r"\[.*\]","",line_re,1)#从左往右替换一次item
                        if line_re == '\n':
                            line_re = line_re.strip()#去除空行
                            continue#跳过本行操作
                        if re.findall(r"\d+、",line) == []:#原本就没写数字信息，直接加
                            line_re = str(task_num)+"、" + line_re
                        task_num += 1
                        work_items[work_item] += line_re
                        #work_items[work_item] += (line + "\n")
                    elif work_item not in line:
                        try:
                            index_flag = line.index('  ')
                        except ValueError as reason:
                            #print(line)
                            #file.remove(line)#删除该行
                            continue#如果不是空格两格，直接跳过
                        if index_flag == 0 and line != '\n':#无标签且前2格为空格，认为是子行，直接加
                            try:
                                son_items.append(line)
                            except AttributeError as reason:
                                print(reason)
                                continue
                            '''#尝试取前一个item，然后添加到对应item下，这样不行，因为读取顺序不对
                            work_item_temp = get_item(file,line)
                            try:
                                if work_items[work_item_temp] == "":
                                    work_items[work_item_temp] += "[%s]"%work_item_temp+"\n"
                                work_items[work_item_temp] += line
                            except KeyError:
                                continue
                            '''
                            file.remove(line)#删除该行
                        else:
                            print(line)
                            work_items['其他'] += line
        with open(reportFile, "w") as fw:#重新写入文档
            fw.write("农信x86组周报\n")
            for work_type, work_contents in work_items.items():
                if work_contents:
                    try:
                        re.findall(r"\[.*\]",work_contents)[0]
                    except IndexError as reason:
                        pass#子项
                    #work_contents = re.sub(r"\[.*\]","",work_contents)
                    fw.write(work_contents)
                fw.write("\n")
        #替换子项
        for son_item in son_items:
            with open(totalFile, "r") as f:#按照标签重新排序
                file = f.readlines()
                #son_item += '\n'
                try:
                    son_item_index = file.index(son_item)#获取子项的index
                except ValueError as reason:
                    print(reason)
                    pass
            item_last_line = file[son_item_index - 1] #获取上一行的内容
            item_last_line = re.sub(r"\[.*\]","",re.sub(r"\d+、","",item_last_line,1),1)#将上一行的内容处理为只要后面的内容,从左往右替换一次数字,从左往右替换一次item
            target_line = item_last_line + son_item#将2行合并作为一行，准备替换
            if item_last_line != None and item_last_line != '\n':
                print(r"原始内容%s\n替换内容%s"%(item_last_line,target_line))
                replace(item_last_line,target_line,reportFile)
            else:
                pass
        os.system("explorer.exe %s" %(reportFile))
        os.chdir(weekReportDir)
        
